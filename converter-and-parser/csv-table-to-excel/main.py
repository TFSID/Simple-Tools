import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

def read_csv_merge_rows(filepath):
    merged_rows = []
    current_row = None

    with open(filepath, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            # Check if the row is mostly empty except first empty cell (common for multiline)
            if current_row and all(not cell.strip() for cell in row[1:]):
                continue  # Skip purely empty rows

            if current_row and not row[1].strip():
                # Merge non-empty cells with previous row
                for i in range(1, len(row)):
                    if row[i].strip():
                        current_row[i] = (current_row[i] + ' ' + row[i]).strip()
            else:
                if current_row:
                    merged_rows.append(current_row)
                current_row = row

    if current_row:
        merged_rows.append(current_row)

    # Remove index column (first column) if it's just empty
    for i, row in enumerate(merged_rows):
        if not row[0].strip():
            merged_rows[i] = row[1:]

    return merged_rows

def create_excel_with_filters(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modules"

    # Write data
    for row in data:
        ws.append(row)

    # Add filter
    col_count = len(data[0])
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}{len(data)}"

    # Adjust column widths
    for col_idx in range(1, col_count + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    wb.save(output_path)

# --- Main ---
csv_file = "sample.csv"             # Input CSV file (replace with your filename)
excel_file = "output.xlsx"          # Output Excel file

data = read_csv_merge_rows(csv_file)
create_excel_with_filters(data, excel_file)

print(f"✅ Excel file with filters saved as: {excel_file}")
